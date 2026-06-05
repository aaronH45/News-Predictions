"""
build_excels.py — produce the De La O & Myers (JF 2021) subjective cash-flow
expectation series in Myers' exact Excel format, extended with two-year columns.

Self-contained: reads the reconstructed series (reconstructed_series.csv, copied
from the WRDS rebuild) + Shiller's long earnings series (for the CAPE denominator),
and writes the two .xlsx files.  See README.md for the full methodology.
"""
import pandas as pd, numpy as np
from pathlib import Path
from openpyxl import Workbook
HERE=Path(__file__).resolve().parent
SHILLER=Path("/home/rpa9/Shiller_div_earnings/ie_data.xls")

# ---- reconstructed series (qe, Earn_ann, Div_ann, SP500, e1, ec2, d1, dc2) ----
A=pd.read_csv(HERE/"reconstructed_series.csv", parse_dates=["qe"]).set_index("qe")

# ---- CAPE denominator e^ca_t from Shiller (10yr cyclically-adjusted earnings) ----
sh=pd.read_excel(SHILLER, sheet_name="Data", header=7)
sh=sh[pd.to_numeric(sh["Date"],errors="coerce").notna()].copy()
yy=sh["Date"].astype(float).astype(int); mm=((sh["Date"].astype(float)-yy)*100).round().astype(int)
sh["dt"]=pd.to_datetime(dict(year=yy,month=mm,day=1),errors="coerce")
sh=sh.dropna(subset=["dt"]).set_index("dt").sort_index()
shq=sh[["E","CPI"]].apply(pd.to_numeric,errors="coerce").resample("QE").last()
# put Shiller earnings on our per-index-point scale (level match over overlap)
ov=A.index.intersection(shq.index)
scale=A.loc[ov,"Earn_ann"].mean()/shq.loc[ov,"E"].mean()
shq["E_s"]=shq.E*scale
shq["real"]=shq.E_s/shq.CPI
shq["cape"]=shq.real.rolling(40,min_periods=40).mean()*shq.CPI    # 10yr (40q) cyc-adj, current $
A["eca"]=np.log(shq["cape"].reindex(A.index))

# ---- levels & shifts ----
A["e_t"]=np.log(A.Earn_ann); A["d_t"]=np.log(A.Div_ann); A["p_t"]=np.log(A.SP500)
A["e_t1"]=np.log(A.Earn_ann.shift(-4)); A["e_t2"]=np.log(A.Earn_ann.shift(-8))
A["d_t1r"]=np.log(A.Div_ann.shift(-4)); A["d_t2r"]=np.log(A.Div_ann.shift(-8))
A["Ee1"]=A.e_t+A.e1; A["Ee2"]=A.e_t+A.ec2        # E*[e_{t+1}], E*[e_{t+2}]
A["Year"]=A.index.year; A["Q"]=A.index.quarter

# ============================ DIVIDEND FILE ============================
dv=A[A.d1.notna()].copy()   # all quarters with dividend forecasts (S&P500 DPS coverage starts 2002)
div=pd.DataFrame({
 "Year":dv.Year,"Quarter":dv.Q,
 "Expected one-year log dividend growth":dv.d1,                        # E*[d_{t+1}-d_t]
 "Realized next year log dividend growth":dv.d_t1r-dv.d_t,            # d_{t+1}-d_t
 "Current log price-dividend ratio":dv.p_t-dv.d_t,                    # p_t-d_t
 "Expected two-year log dividend growth":dv.dc2,                      # E*[d_{t+2}-d_t]
 "Realized two-year log dividend growth":dv.d_t2r-dv.d_t,            # d_{t+2}-d_t
}).reset_index(drop=True)
div.to_excel(HERE/"Dividend_growth_expectations_reconstructed.xlsx",index=False)

# ============================ EARNINGS FILE ============================
en=A[A.e1.notna()].copy()   # all quarters with earnings forecasts (EPS coverage from 1976)
def block(x):   # x = denominator series (e_t, d_t, or eca)
    return pd.DataFrame({
      "Expected one-year log earnings growth": en.Ee1-x,
      "Realized next year log earnings growth": en.e_t1-x,
      "Current price ratio": en.p_t-x,
      "Expected two-year log earnings growth": en.Ee2-x,
      "Realized two-year log earnings growth": en.e_t2-x})
B=[("Denominator is current earnings e_t",block(en.e_t)),
   ("Denominator is dividends d_t",block(en.d_t)),
   ("Denominator is 10yr cyclically adjusted earnings from CAPE e^ca_t",block(en.eca))]
wb=Workbook(); ws=wb.active; ws.title="Sheet1"
ws.cell(row=2,column=1,value="Year"); ws.cell(row=2,column=2,value="Quarter")
col=3
for label,blk in B:
    ws.cell(row=1,column=col,value=label)
    for j,cn in enumerate(blk.columns): ws.cell(row=2,column=col+j,value=cn)
    col+=len(blk.columns)
for i,(_,row) in enumerate(en.iterrows()):
    r=i+3; ws.cell(row=r,column=1,value=int(row.Year)); ws.cell(row=r,column=2,value=int(row.Q))
    c=3
    for _,blk in B:
        for cn in blk.columns:
            v=blk.iloc[i][cn]; ws.cell(row=r,column=c,value=None if pd.isna(v) else float(v)); c+=1
wb.save(HERE/"Earnings_growth_expectations_reconstructed.xlsx")
drange=f"{int(div.Year.iloc[0])}Q{int(div.Quarter.iloc[0])}-{int(div.Year.iloc[-1])}Q{int(div.Quarter.iloc[-1])}"
erange=f"{int(en.Year.iloc[0])}Q{int(en.Q.iloc[0])}-{int(en.Year.iloc[-1])}Q{int(en.Q.iloc[-1])}"
print(f"wrote dividend file: {div.shape}  ({drange})")
print(f"wrote earnings file: {en.shape[0]} rows x 17 cols ({erange}; 2yr cols start where FY3 exists)")
print(f"\nCAPE scale factor (Shiller E -> our level): {scale:.3f}")
print("sanity — log CAPE ratio (p_t - e^ca_t) should ~ log(Shiller CAPE):")
print((A.p_t-A.eca).dropna().loc[["2015-03-31","2021-09-30"]].round(3).to_string())
print("\nDIVIDEND head:"); print(div.head(2).to_string(index=False))
print("\nEARNINGS sample (e-denominator block, 1976 & 2015):")
print(en[en.Year.isin([1976,2015])&(en.Q==1)][["Year","Q","e1","ec2"]].round(4).to_string(index=False))
